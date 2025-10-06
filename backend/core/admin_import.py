"""
Wizard импорта промокодов из CSV/XLSX
Поддерживает загрузку, маппинг колонок, превью и валидацию
"""
import csv
import logging
from datetime import datetime
from io import StringIO

from django.contrib import admin, messages
from django.http import HttpResponse, Http404
from django.shortcuts import render, redirect
from django.urls import path
from django.utils.html import format_html

logger = logging.getLogger(__name__)


def parse_date_flexible(date_str):
    """Парсинг даты в различных форматах"""
    if not date_str:
        return None

    date_str = date_str.strip()

    # Форматы для попытки парсинга
    formats = [
        '%Y-%m-%d',      # 2025-12-31
        '%d.%m.%Y',      # 31.12.2025
        '%d/%m/%Y',      # 31/12/2025
        '%Y/%m/%d',      # 2025/12/31
        '%d-%m-%Y',      # 31-12-2025
    ]

    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue

    # Если не получилось - пробуем dateutil если доступен
    try:
        from dateutil.parser import parse
        return parse(date_str, dayfirst=True)
    except:
        return None


def download_template(request):
    """Скачать шаблон CSV для импорта"""
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="promo_template.csv"'

    writer = csv.writer(response)
    writer.writerow([
        'title', 'description', 'store_slug', 'discount_value', 'discount_label',
        'offer_type', 'code', 'expires_at', 'url', 'is_active', 'is_hot', 'is_recommended'
    ])
    writer.writerow([
        'Пример: Скидка 20% на все',
        'Описание промокода',
        'ozon',  # slug магазина
        '20',
        '20% скидка',
        'coupon',  # coupon/deal/financial/cashback
        'PROMO20',
        '2025-12-31',
        'https://example.com/promo',
        'true',
        'false',
        'false'
    ])

    return response


def import_promocodes_view(request):
    """Основной view для импорта промокодов"""
    from .models import Store, PromoCode
    import re

    if not request.user.has_perm('core.add_promocode'):
        raise Http404("У вас нет прав для импорта промокодов")

    # Шаг 1: Загрузка файла
    if request.method == 'GET':
        stores = Store.objects.filter(is_active=True).order_by('name')
        return render(request, 'admin/import_step1.html', {
            'title': 'Импорт промокодов - Шаг 1: Загрузка файла',
            'stores': stores,
            'import_profiles': [
                {'value': 'standard', 'label': 'Стандартный (12 полей)'},
                {'value': 'partner', 'label': 'Партнёрский CSV (4 поля: store_name, offer, conditions, date)'},
            ]
        })

    # Шаг 2: Обработка файла
    if request.method == 'POST':
        uploaded_file = request.FILES.get('file')
        import_profile = request.POST.get('import_profile', 'standard')

        if not uploaded_file:
            messages.error(request, 'Файл не загружен')
            return redirect('admin:import_promocodes')

        # Проверка расширения
        filename = uploaded_file.name.lower()
        if not (filename.endswith('.csv') or filename.endswith('.xlsx')):
            messages.error(request, 'Поддерживаются только CSV и XLSX файлы')
            return redirect('admin:import_promocodes')

        try:
            # Парсинг CSV
            if filename.endswith('.csv'):
                decoded_file = uploaded_file.read().decode('utf-8-sig')
                io_string = StringIO(decoded_file)
                reader = csv.DictReader(io_string)
                rows = list(reader)

            # Парсинг XLSX
            elif filename.endswith('.xlsx'):
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(uploaded_file, read_only=True)
                    ws = wb.active
                    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                    rows = []
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        rows.append(dict(zip(headers, row)))
                except ImportError:
                    messages.error(request, 'Для импорта XLSX установите библиотеку openpyxl: pip install openpyxl')
                    return redirect('admin:import_promocodes')

            # Валидация и превью
            if not rows:
                messages.error(request, 'Файл пуст или не содержит данных')
                return redirect('admin:import_promocodes')

            # Валидация каждой строки
            validated_rows = []
            errors = []

            for idx, row in enumerate(rows, start=2):  # начинаем с 2 т.к. первая строка - заголовки
                # Партнёрский CSV профиль
                if import_profile == 'partner':
                    store_name = row.get('store_name', '').strip()
                    offer_text = row.get('offer', '').strip()
                    conditions = row.get('conditions', '').strip()
                    date_text = row.get('date', '').strip()

                    validated_row = {
                        'line': idx,
                        'profile': 'partner',
                        'store_name': store_name,
                        'offer': offer_text,
                        'conditions': conditions,
                        'date': date_text,
                        'errors': [],
                        'warnings': [],
                        'valid': True,
                        # Обработанные поля (будут заполнены ниже)
                        'store_slug': '',
                        'title': '',
                        'description': '',
                        'offer_type': '',
                        'code': '',
                        'fine_print': conditions,
                        'expires_at': date_text,
                        'affiliate_url': '',
                        'is_active': True,
                        'is_hot': False,
                        'is_recommended': False,
                    }

                    # 1. Определение магазина
                    if not store_name:
                        validated_row['errors'].append('Отсутствует название магазина')
                        validated_row['valid'] = False
                    else:
                        # Поиск магазина (case-insensitive)
                        store = Store.objects.filter(name__iexact=store_name).first()
                        if not store:
                            validated_row['errors'].append(f'Магазин "{store_name}" не найден. Создайте магазин перед импортом.')
                            validated_row['valid'] = False
                        else:
                            validated_row['store_slug'] = store.slug

                    # 2. Обработка поля offer (промокод или текст предложения)
                    if not offer_text:
                        validated_row['errors'].append('Отсутствует предложение/промокод')
                        validated_row['valid'] = False
                    else:
                        # Проверка: это промокод или текстовое предложение?
                        if re.match(r'^[A-Z0-9-]{4,}$', offer_text):
                            # Это промокод
                            validated_row['offer_type'] = 'coupon'
                            validated_row['code'] = offer_text
                            validated_row['title'] = f'Промокод {offer_text}'
                            validated_row['description'] = conditions or f'Используйте промокод {offer_text}'
                            validated_row['warnings'].append('⚠️ affiliate_url пустой - потребуется редактирование')
                        else:
                            # Это текстовое предложение
                            validated_row['title'] = offer_text[:200]  # Ограничиваем длину
                            validated_row['description'] = conditions or offer_text

                            # Определяем тип по ключевым словам
                            offer_lower = offer_text.lower()
                            if any(word in offer_lower for word in ['кредит', 'банк', 'кэшбэк', 'cashback', 'займ', 'карта']):
                                validated_row['offer_type'] = 'financial'
                            elif any(word in offer_lower for word in ['скидка', 'распродажа', '%', 'процент', 'sale', 'discount']):
                                validated_row['offer_type'] = 'deal'
                            else:
                                validated_row['offer_type'] = 'deal'  # По умолчанию

                            validated_row['warnings'].append('⚠️ affiliate_url пустой - потребуется редактирование')

                    # 3. Обработка даты (expires_at)
                    if date_text:
                        # Пробуем распознать формат
                        validated_row['expires_at'] = date_text  # Сохраняем как есть, парсинг позже
                    else:
                        validated_row['warnings'].append('⚠️ Дата отсутствует')

                # Стандартный профиль
                else:
                    validated_row = {
                        'line': idx,
                        'profile': 'standard',
                        'title': row.get('title', '').strip(),
                        'description': row.get('description', '').strip(),
                        'store_slug': row.get('store_slug', '').strip(),
                        'discount_value': row.get('discount_value', '').strip(),
                        'discount_label': row.get('discount_label', '').strip(),
                        'offer_type': row.get('offer_type', 'coupon').strip(),
                        'code': row.get('code', '').strip(),
                        'expires_at': row.get('expires_at', '').strip(),
                        'affiliate_url': row.get('url', '').strip(),
                        'is_active': str(row.get('is_active', 'true')).lower() in ('true', '1', 'yes'),
                        'is_hot': str(row.get('is_hot', 'false')).lower() in ('true', '1', 'yes'),
                        'is_recommended': str(row.get('is_recommended', 'false')).lower() in ('true', '1', 'yes'),
                        'errors': [],
                        'warnings': [],
                        'valid': True
                    }

                    # Валидация для стандартного профиля
                    if not validated_row['title']:
                        validated_row['errors'].append('Отсутствует название')
                        validated_row['valid'] = False

                    if not validated_row['store_slug']:
                        validated_row['errors'].append('Отсутствует slug магазина')
                        validated_row['valid'] = False
                    else:
                        # Проверка существования магазина
                        if not Store.objects.filter(slug=validated_row['store_slug']).exists():
                            validated_row['errors'].append(f'Магазин с slug "{validated_row["store_slug"]}" не найден')
                            validated_row['valid'] = False

                    if validated_row['offer_type'] not in ['coupon', 'deal', 'financial', 'cashback']:
                        validated_row['errors'].append(f'Неверный тип оффера: {validated_row["offer_type"]}')
                        validated_row['valid'] = False

                validated_rows.append(validated_row)

            # Подсчет статистики
            valid_count = sum(1 for r in validated_rows if r['valid'])
            invalid_count = len(validated_rows) - valid_count

            # Превью
            request.session['import_data'] = {
                'rows': validated_rows,
                'total': len(validated_rows),
                'valid': valid_count,
                'invalid': invalid_count
            }

            return render(request, 'admin/import_step2.html', {
                'title': 'Импорт промокодов - Шаг 2: Превью и валидация',
                'rows': validated_rows[:50],  # Показываем первые 50
                'total': len(validated_rows),
                'valid_count': valid_count,
                'invalid_count': invalid_count,
            })

        except Exception as e:
            logger.error(f"Import error: {str(e)}", exc_info=True)
            messages.error(request, f'Ошибка обработки файла: {str(e)}')
            return redirect('admin:import_promocodes')

    return redirect('admin:import_promocodes')


def import_execute_view(request):
    """Выполнение импорта"""
    from .models import Store, PromoCode
    from datetime import datetime

    if not request.user.has_perm('core.add_promocode'):
        raise Http404("У вас нет прав для импорта промокодов")

    if request.method != 'POST':
        return redirect('admin:import_promocodes')

    import_data = request.session.get('import_data')
    if not import_data:
        messages.error(request, 'Данные импорта не найдены. Загрузите файл заново.')
        return redirect('admin:import_promocodes')

    rows = import_data['rows']

    # Импорт только валидных строк
    imported = 0
    skipped = 0
    errors = []

    for row in rows:
        if not row['valid']:
            skipped += 1
            continue

        try:
            store = Store.objects.get(slug=row['store_slug'])

            # Парсинг даты
            expires_at = None
            if row['expires_at']:
                try:
                    expires_at = datetime.strptime(row['expires_at'], '%Y-%m-%d')
                except ValueError:
                    try:
                        # Попытка парсинга без времени
                        from dateutil.parser import parse
                        expires_at = parse(row['expires_at'])
                    except:
                        expires_at = None

            # Создание промокода с ignore_conflicts
            PromoCode.objects.get_or_create(
                title=row['title'],
                store=store,
                defaults={
                    'description': row['description'],
                    'discount_value': int(row['discount_value']) if row['discount_value'] else 0,
                    'discount_label': row['discount_label'],
                    'offer_type': row['offer_type'],
                    'code': row['code'],
                    'expires_at': expires_at,
                    'affiliate_url': row['affiliate_url'],
                    'is_active': row['is_active'],
                    'is_hot': row['is_hot'],
                    'is_recommended': row['is_recommended'],
                }
            )

            imported += 1

        except Exception as e:
            logger.error(f"Error importing row {row['line']}: {str(e)}")
            errors.append(f"Строка {row['line']}: {str(e)}")
            skipped += 1

    # Очистка сессии
    if 'import_data' in request.session:
        del request.session['import_data']

    # Сообщения
    if imported > 0:
        messages.success(request, f'Успешно импортировано промокодов: {imported}')

    if skipped > 0:
        messages.warning(request, f'Пропущено строк: {skipped}')

    if errors:
        messages.error(request, f'Ошибки импорта:\n' + '\n'.join(errors[:10]))

    return redirect('admin:core_promocode_changelist')


# URL patterns для админки
def get_import_urls():
    """Возвращает URL patterns для импорта"""
    return [
        path('import/', import_promocodes_view, name='import_promocodes'),
        path('import/execute/', import_execute_view, name='import_execute'),
        path('import/template/', download_template, name='import_template'),
    ]
